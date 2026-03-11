#!/usr/bin/env python3
"""
Generate white space matrix visualizations from the scored database.

Produces:
1. Bubble chart: X=Drug Coverage Gap, Y=Unmet Need, Size=Prevalence, Color=Complexity
2. Bubble chart (health econ variant): Color by health_econ_score instead of complexity
3. Ranked bar chart: Top 20 diseases by composite score
4. Therapeutic area summary chart
5. Specialty segment comparison: Grouped bar chart of pediatric, geriatric, rare, general segments
6. Regulatory landscape: Horizontal bar chart sorted by regulatory advantage

Usage:
    python visualize_matrix.py <path_to_workbook> [--output-dir <path>]
"""

import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.cm as cm
    from matplotlib.patches import FancyBboxPatch
    import numpy as np
except ImportError:
    os.system(f"{sys.executable} -m pip install matplotlib numpy --break-system-packages -q")
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.cm as cm
    from matplotlib.patches import FancyBboxPatch
    import numpy as np


def read_scores(wb):
    """Read White Space Scores tab into list of dicts."""
    ws = wb["White Space Scores"]
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value.lower().replace(" ", "_"))
        else:
            headers.append(None)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[0] == "":
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val
        rows.append(d)
    return rows


def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def bubble_chart(scores, output_path, color_by="scientific_complexity", variant_name=""):
    """Create the main white space bubble chart.

    Args:
        scores: List of disease score dicts
        output_path: Path to save the figure
        color_by: Column to use for coloring bubbles (default: scientific_complexity)
        variant_name: Suffix for output filename (e.g., "_health_econ"), empty for main
    """
    fig, ax = plt.subplots(figsize=(16, 11))

    x_vals = [safe_float(s.get("drug_coverage_gap")) for s in scores]
    y_vals = [safe_float(s.get("unmet_need_score")) for s in scores]
    sizes = [max(safe_float(s.get("us_prevalence")) / 50000, 20) for s in scores]
    # Cap bubble size for readability
    sizes = [min(s, 800) for s in sizes]
    colors = [safe_float(s.get(color_by)) for s in scores]
    names = [s.get("disease_name", "Unknown") for s in scores]

    # Normalize colors to 0-10 range for colormap
    norm = plt.Normalize(vmin=0, vmax=10)
    cmap = cm.RdYlGn_r  # Red=complex/low, Green=simpler/high

    scatter = ax.scatter(x_vals, y_vals, s=sizes, c=colors,
                        cmap=cmap, norm=norm, alpha=0.7,
                        edgecolors="white", linewidth=0.5)

    # Label top opportunities (high composite score)
    sorted_scores = sorted(scores,
                          key=lambda s: safe_float(s.get("composite_whitespace_score")),
                          reverse=True)
    top_n = min(15, len(sorted_scores))
    top_names = set(s.get("disease_name") for s in sorted_scores[:top_n])

    for i, name in enumerate(names):
        if name in top_names:
            ax.annotate(name, (x_vals[i], y_vals[i]),
                       fontsize=7, ha="center", va="bottom",
                       xytext=(0, 6), textcoords="offset points",
                       fontweight="bold", color="#333333",
                       bbox=dict(boxstyle="round,pad=0.2",
                                facecolor="white", alpha=0.8, edgecolor="none"))

    # Quadrant lines and labels
    ax.axhline(y=5, color="#cccccc", linestyle="--", linewidth=0.8)
    ax.axvline(x=5, color="#cccccc", linestyle="--", linewidth=0.8)

    ax.text(8, 9.5, "HIGH OPPORTUNITY\nHigh Need + Few Drugs",
            ha="center", va="top", fontsize=9, color="#2F5496",
            fontweight="bold", alpha=0.6)
    ax.text(2, 9.5, "ACTIVE INVESTMENT\nHigh Need + Many Drugs",
            ha="center", va="top", fontsize=9, color="#C55A11",
            fontweight="bold", alpha=0.6)
    ax.text(8, 0.5, "NICHE OPPORTUNITY\nLow Need + Few Drugs",
            ha="center", va="bottom", fontsize=9, color="#548235",
            fontweight="bold", alpha=0.6)
    ax.text(2, 0.5, "WELL-SERVED\nLow Need + Many Drugs",
            ha="center", va="bottom", fontsize=9, color="#808080",
            fontweight="bold", alpha=0.6)

    # Axes
    ax.set_xlabel("Drug Coverage Gap (higher = fewer drugs)", fontsize=12, fontweight="bold")
    ax.set_ylabel("Unmet Patient Need (higher = more need)", fontsize=12, fontweight="bold")

    # Determine title and colorbar label based on color dimension
    if color_by == "health_econ_score":
        title = "Drug Development White Space Matrix (Health Economics View)"
        cbar_label = "Health Economics Score\n(Red=Low, Green=High)"
    else:
        title = "Drug Development White Space Matrix"
        cbar_label = "Scientific Complexity\n(Red=Complex, Green=Simpler)"

    ax.set_title(title, fontsize=16, fontweight="bold", pad=20)
    ax.set_xlim(-0.5, 10.5)
    ax.set_ylim(-0.5, 10.5)
    ax.set_xticks(range(0, 11))
    ax.set_yticks(range(0, 11))
    ax.grid(True, alpha=0.2)

    # Colorbar
    cbar = plt.colorbar(scatter, ax=ax, shrink=0.6, pad=0.02)
    cbar.set_label(cbar_label, fontsize=10, fontweight="bold")

    # Size legend
    legend_sizes = [1000, 10000000, 30000000]
    legend_labels = ["1K patients", "10M patients", "30M patients"]
    for sz, label in zip(legend_sizes, legend_labels):
        bubble_sz = min(max(sz / 50000, 20), 800)
        ax.scatter([], [], s=bubble_sz, c="gray", alpha=0.4,
                  edgecolors="white", label=label)
    ax.legend(loc="lower left", title="US Prevalence", fontsize=8,
             title_fontsize=9, framealpha=0.9)

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"Saved bubble chart: {output_path}")


def ranked_bar_chart(scores, output_path, top_n=20):
    """Create a horizontal bar chart of top diseases by composite score."""
    sorted_scores = sorted(scores,
                          key=lambda s: safe_float(s.get("composite_whitespace_score")),
                          reverse=True)[:top_n]

    fig, ax = plt.subplots(figsize=(12, max(8, top_n * 0.4)))

    names = [s.get("disease_name", "Unknown") for s in reversed(sorted_scores)]
    composites = [safe_float(s.get("composite_whitespace_score")) for s in reversed(sorted_scores)]
    unmet = [safe_float(s.get("unmet_need_score")) for s in reversed(sorted_scores)]
    gaps = [safe_float(s.get("drug_coverage_gap")) for s in reversed(sorted_scores)]

    y_pos = np.arange(len(names))

    # Stacked bars showing component contributions
    ax.barh(y_pos, [u * W_UNMET for u in unmet], height=0.6,
            label="Unmet Need (40%)", color="#2F5496", alpha=0.85)
    ax.barh(y_pos, [g * W_GAP for g in gaps], height=0.6,
            left=[u * W_UNMET for u in unmet],
            label="Coverage Gap (35%)", color="#C55A11", alpha=0.85)
    feasibility = [c - (u * W_UNMET + g * W_GAP)
                  for c, u, g in zip(composites, unmet, gaps)]
    ax.barh(y_pos, feasibility, height=0.6,
            left=[u * W_UNMET + g * W_GAP for u, g in zip(unmet, gaps)],
            label="Feasibility (25%)", color="#548235", alpha=0.85)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(names, fontsize=9)
    ax.set_xlabel("Composite White Space Score", fontsize=12, fontweight="bold")
    ax.set_title(f"Top {top_n} Drug Development Opportunities",
                fontsize=14, fontweight="bold")
    ax.set_xlim(0, 10)
    ax.legend(loc="lower right", fontsize=9)
    ax.grid(axis="x", alpha=0.2)

    # Score labels
    for i, v in enumerate(composites):
        ax.text(v + 0.1, i, f"{v:.1f}", va="center", fontsize=8, fontweight="bold")

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"Saved ranked chart: {output_path}")


# Component weights for stacked bar (must match scoring_methodology.md)
W_UNMET = 0.40
W_GAP = 0.35


def therapeutic_area_summary(scores, output_path):
    """Create a summary chart by therapeutic area."""
    from collections import defaultdict

    area_data = defaultdict(lambda: {"count": 0, "avg_score": 0, "total_prevalence": 0,
                                      "avg_gap": 0, "scores": []})

    for s in scores:
        area = s.get("therapeutic_area", "Unknown")
        composite = safe_float(s.get("composite_whitespace_score"))
        prevalence = safe_float(s.get("us_prevalence"))
        gap = safe_float(s.get("drug_coverage_gap"))

        area_data[area]["count"] += 1
        area_data[area]["total_prevalence"] += prevalence
        area_data[area]["avg_gap"] += gap
        area_data[area]["scores"].append(composite)

    # Compute averages
    for area in area_data:
        n = area_data[area]["count"]
        area_data[area]["avg_score"] = sum(area_data[area]["scores"]) / n
        area_data[area]["avg_gap"] = area_data[area]["avg_gap"] / n

    # Sort by avg score
    sorted_areas = sorted(area_data.items(), key=lambda x: x[1]["avg_score"], reverse=True)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, max(6, len(sorted_areas) * 0.5)))

    areas = [a[0] for a in reversed(sorted_areas)]
    avg_scores = [a[1]["avg_score"] for a in reversed(sorted_areas)]
    counts = [a[1]["count"] for a in reversed(sorted_areas)]
    avg_gaps = [a[1]["avg_gap"] for a in reversed(sorted_areas)]

    y_pos = np.arange(len(areas))

    # Left: Avg composite score
    colors = ["#2F5496" if s >= 6 else "#C55A11" if s >= 4 else "#808080" for s in avg_scores]
    ax1.barh(y_pos, avg_scores, color=colors, alpha=0.8, height=0.6)
    ax1.set_yticks(y_pos)
    ax1.set_yticklabels(areas, fontsize=9)
    ax1.set_xlabel("Avg Composite Score", fontsize=11, fontweight="bold")
    ax1.set_title("Avg White Space Score by Area", fontsize=13, fontweight="bold")
    ax1.set_xlim(0, 10)
    for i, (v, c) in enumerate(zip(avg_scores, counts)):
        ax1.text(v + 0.1, i, f"{v:.1f} ({c} diseases)", va="center", fontsize=8)

    # Right: Avg drug coverage gap
    ax2.barh(y_pos, avg_gaps, color="#C55A11", alpha=0.7, height=0.6)
    ax2.set_yticks(y_pos)
    ax2.set_yticklabels([""] * len(areas))
    ax2.set_xlabel("Avg Drug Coverage Gap", fontsize=11, fontweight="bold")
    ax2.set_title("Avg Coverage Gap by Area", fontsize=13, fontweight="bold")
    ax2.set_xlim(0, 10)
    for i, v in enumerate(avg_gaps):
        ax2.text(v + 0.1, i, f"{v:.1f}", va="center", fontsize=8)

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"Saved area summary: {output_path}")


def specialty_segment_chart(scores, output_path):
    """Create grouped bar chart comparing specialty segments.

    Segments: pediatric, geriatric, rare, general
    For each segment: count, avg composite score, avg coverage gap
    """
    from collections import defaultdict

    # Define segment keywords for classification
    segment_keywords = {
        "pediatric": ["pediatric", "child", "infant", "neonatal", "congenital"],
        "geriatric": ["aging", "elderly", "geriatric", "dementia", "alzheimer"],
        "rare": ["rare", "orphan", "ultra-rare"],
        "general": []
    }

    segment_data = {seg: {"count": 0, "avg_score": 0, "avg_gap": 0, "scores": [], "gaps": []}
                   for seg in segment_keywords.keys()}

    # Classify diseases into segments based on disease_name
    for s in scores:
        disease_name = s.get("disease_name", "").lower()
        segment = "general"  # default

        for seg_type, keywords in segment_keywords.items():
            if seg_type != "general" and any(kw in disease_name for kw in keywords):
                segment = seg_type
                break

        composite = safe_float(s.get("composite_whitespace_score"))
        gap = safe_float(s.get("drug_coverage_gap"))

        segment_data[segment]["count"] += 1
        segment_data[segment]["scores"].append(composite)
        segment_data[segment]["gaps"].append(gap)

    # Compute averages
    for segment in segment_data:
        if segment_data[segment]["count"] > 0:
            segment_data[segment]["avg_score"] = sum(segment_data[segment]["scores"]) / segment_data[segment]["count"]
            segment_data[segment]["avg_gap"] = sum(segment_data[segment]["gaps"]) / segment_data[segment]["count"]

    # Sort segments in a logical order
    segment_order = ["pediatric", "geriatric", "rare", "general"]
    segments = [s for s in segment_order if segment_data[s]["count"] > 0]

    # Prepare data for grouped bar chart
    counts = [segment_data[seg]["count"] for seg in segments]
    avg_scores = [segment_data[seg]["avg_score"] for seg in segments]
    avg_gaps = [segment_data[seg]["avg_gap"] for seg in segments]

    # Color mapping
    segment_colors = {
        "pediatric": "#1f77b4",  # blue
        "geriatric": "#ff7f0e",  # orange
        "rare": "#9467bd",       # purple
        "general": "#7f7f7f"     # gray
    }
    colors = [segment_colors[seg] for seg in segments]

    fig, ax = plt.subplots(figsize=(14, 7))

    x_pos = np.arange(len(segments))
    width = 0.25

    # Create grouped bars
    bars1 = ax.bar(x_pos - width, counts, width, label="Number of Diseases",
                   color=colors, alpha=0.7)
    ax2 = ax.twinx()
    bars2 = ax2.bar(x_pos, avg_scores, width, label="Avg Composite Score",
                    color=colors, alpha=0.9, edgecolor="black", linewidth=1.5)
    bars3 = ax2.bar(x_pos + width, avg_gaps, width, label="Avg Coverage Gap",
                    color=colors, alpha=0.5, hatch="//")

    ax.set_xlabel("Specialty Segment", fontsize=12, fontweight="bold")
    ax.set_ylabel("Number of Diseases", fontsize=11, fontweight="bold", color="#333333")
    ax2.set_ylabel("Score (0-10)", fontsize=11, fontweight="bold", color="#333333")
    ax.set_title("White Space by Specialty Segment", fontsize=14, fontweight="bold", pad=20)

    ax.set_xticks(x_pos)
    ax.set_xticklabels(segments, fontsize=11, fontweight="bold")
    ax.set_ylim(0, max(counts) * 1.2)
    ax2.set_ylim(0, 10)
    ax.grid(axis="y", alpha=0.3)

    # Add value labels on bars
    for bars in [bars1]:
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width() / 2., height,
                       f"{int(height)}", ha="center", va="bottom", fontsize=9, fontweight="bold")

    for bars in [bars2, bars3]:
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax2.text(bar.get_x() + bar.get_width() / 2., height,
                        f"{height:.1f}", ha="center", va="bottom", fontsize=8)

    # Combined legend
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=10)

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"Saved specialty segment chart: {output_path}")


def regulatory_landscape_chart(scores, output_path):
    """Create horizontal bar chart sorted by regulatory advantage score.

    Colors bars by composite whitespace score (RdYlGn colormap).
    Annotates with count of orphan + breakthrough designations.
    Shows top 20 if more than 20 diseases.
    """
    # Check if regulatory_advantage_score column exists
    has_regulatory = any(s.get("regulatory_advantage_score") is not None for s in scores)

    if not has_regulatory:
        print(f"Skipping regulatory landscape chart: regulatory_advantage_score column not found")
        return

    # Prepare data
    diseases_with_regulatory = [s for s in scores if s.get("regulatory_advantage_score") is not None]

    if not diseases_with_regulatory:
        print(f"Skipping regulatory landscape chart: no regulatory data available")
        return

    # Sort by regulatory advantage
    sorted_diseases = sorted(diseases_with_regulatory,
                            key=lambda s: safe_float(s.get("regulatory_advantage_score")),
                            reverse=True)

    # Limit to top 20
    if len(sorted_diseases) > 20:
        sorted_diseases = sorted_diseases[:20]

    names = [s.get("disease_name", "Unknown") for s in sorted_diseases]
    reg_scores = [safe_float(s.get("regulatory_advantage_score")) for s in sorted_diseases]
    composite_scores = [safe_float(s.get("composite_whitespace_score")) for s in sorted_diseases]

    # Count orphan + breakthrough designations (sum of two columns if they exist)
    designation_counts = []
    for s in sorted_diseases:
        orphan = safe_float(s.get("orphan_designation"), 0.0)
        breakthrough = safe_float(s.get("breakthrough_designation"), 0.0)
        designation_counts.append(int(orphan + breakthrough))

    fig, ax = plt.subplots(figsize=(14, max(8, len(sorted_diseases) * 0.35)))

    y_pos = np.arange(len(names))

    # Color bars by composite whitespace score using RdYlGn
    norm = plt.Normalize(vmin=0, vmax=10)
    cmap = cm.RdYlGn

    colors = [cmap(norm(cs)) for cs in composite_scores]

    bars = ax.barh(y_pos, reg_scores, color=colors, alpha=0.85, edgecolor="black", linewidth=0.5)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(names, fontsize=8)
    ax.set_xlabel("Regulatory Advantage Score", fontsize=12, fontweight="bold")
    ax.set_title("Regulatory Advantage by Disease", fontsize=14, fontweight="bold", pad=20)
    ax.grid(axis="x", alpha=0.3)

    # Annotate bars with designation counts
    for i, (bar, count) in enumerate(zip(bars, designation_counts)):
        width = bar.get_width()
        ax.text(width + 0.15, bar.get_y() + bar.get_height() / 2.,
               f"({count} desig.)", ha="left", va="center", fontsize=7, style="italic")

    # Add colorbar to show composite whitespace score mapping
    sm = cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, pad=0.02, shrink=0.8)
    cbar.set_label("Composite Whitespace Score", fontsize=10, fontweight="bold")

    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"Saved regulatory landscape chart: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Visualize drug whitespace matrix")
    parser.add_argument("workbook_path", help="Path to the Excel workbook")
    parser.add_argument("--output-dir", default=None, help="Directory for output images")
    args = parser.parse_args()

    if not os.path.exists(args.workbook_path):
        print(f"Error: Workbook not found: {args.workbook_path}")
        sys.exit(1)

    output_dir = args.output_dir or os.path.dirname(args.workbook_path) or "."
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(args.workbook_path)
    scores = read_scores(wb)

    if not scores:
        print("No scores found. Run score_diseases.py first.")
        sys.exit(1)

    print(f"Generating visualizations for {len(scores)} diseases...")

    # Generate primary bubble chart (complexity-based)
    bubble_chart(scores, os.path.join(output_dir, "whitespace_bubble_chart.png"),
                color_by="scientific_complexity")

    # Generate health economics variant if data exists
    has_health_econ = any(s.get("health_econ_score") is not None for s in scores)
    if has_health_econ:
        bubble_chart(scores, os.path.join(output_dir, "whitespace_bubble_health_econ.png"),
                    color_by="health_econ_score")
        print("Generated health economics bubble chart variant")

    # Generate other visualizations
    ranked_bar_chart(scores, os.path.join(output_dir, "whitespace_ranked_chart.png"))

    if len(set(s.get("therapeutic_area", "") for s in scores)) > 1:
        therapeutic_area_summary(scores, os.path.join(output_dir, "whitespace_area_summary.png"))

    # Generate specialty segment chart
    specialty_segment_chart(scores, os.path.join(output_dir, "whitespace_specialty_segments.png"))

    # Generate regulatory landscape chart
    regulatory_landscape_chart(scores, os.path.join(output_dir, "whitespace_regulatory_landscape.png"))

    print(f"\nAll visualizations saved to: {output_dir}")


if __name__ == "__main__":
    main()

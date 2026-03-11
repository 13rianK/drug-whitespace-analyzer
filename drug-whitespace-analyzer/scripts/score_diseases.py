#!/usr/bin/env python3
"""
Score diseases in the whitespace database and write results to the White Space Scores tab.

v2.0 Scoring includes:
- Regulatory Advantage Score (orphan drug, breakthrough therapy, fast track designations)
- Health Economics Score (cost per patient, economic burden, QALY burden)
- Approval Probability modeling based on pipeline composition
- Historical trend tracking

Usage:
    python score_diseases.py <path_to_workbook>

Reads from: Disease Epidemiology, Approved Drugs, Pipeline Drugs, Basic Science Activity, Approval Probability (optional)
Writes to:  White Space Scores tab, Historical Trends tab
"""

import argparse
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side


# ── Scoring Parameters ────────────────────────────────────────────────

SEVERITY_WEIGHTS = {
    "life-threatening": 1.0,
    "debilitating": 0.75,
    "chronic-manageable": 0.5,
    "acute-treatable": 0.25,
}

PHASE_WEIGHTS = {
    "PHASE1": 0.1,
    "EARLY_PHASE1": 0.1,
    "PHASE2": 0.3,
    "PHASE3": 0.6,
    "PHASE4": 0.9,
}

# Approval probability phase weights for expected approvals
APPROVAL_PHASE_WEIGHTS = {
    "PHASE1": 0.05,
    "EARLY_PHASE1": 0.05,
    "PHASE2": 0.15,
    "PHASE3": 0.40,
    "PHASE4": 0.80,
}

# v2.0 Composite score weights
W_UNMET_NEED = 0.30
W_COVERAGE_GAP = 0.25
W_COMPLEXITY_INV = 0.15  # inverted: lower complexity = higher score contribution
W_REGULATORY_ADV = 0.15
W_HEALTH_ECON = 0.15


def normalize_log(value, min_val, max_val):
    """Log-normalize a value to [0, 1] range."""
    if value <= min_val:
        return 0.0
    if value >= max_val:
        return 1.0
    return math.log(value / min_val) / math.log(max_val / min_val)


def clamp(value, low=0.0, high=10.0):
    return max(low, min(high, value))


def read_tab_as_dicts(ws):
    """Read an openpyxl worksheet into a list of dicts using header row as keys."""
    headers = []
    for cell in ws[1]:
        if cell.value:
            # Convert "Disease Name" back to "disease_name"
            headers.append(cell.value.lower().replace(" ", "_"))
        else:
            headers.append(None)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[0] == "":
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        rows.append(row_dict)
    return rows


def safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_bool(val, default=False):
    """Convert various boolean representations to bool."""
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() in ("yes", "true", "1", "y", "t")
    return bool(val)


def build_approval_probability_lookup(wb):
    """
    Build a lookup of approval probabilities by therapeutic area and phase.
    Returns a dict: {therapeutic_area: {phase: probability}} or empty dict if tab not found.
    """
    lookup = {}
    try:
        ws = wb["Approval Probability"]
        data = read_tab_as_dicts(ws)
        for row in data:
            ta = row.get("therapeutic_area", "").strip()
            phase = row.get("trial_phase", "").strip()
            prob = safe_float(row.get("approval_probability"), 0.5)
            if ta and phase:
                if ta not in lookup:
                    lookup[ta] = {}
                lookup[ta][phase] = prob
    except KeyError:
        # Approval Probability tab not found; use defaults
        pass
    return lookup


def compute_scores(epi_data, drugs_data, pipeline_data, science_data, approval_prob_lookup=None):
    """Compute white space scores for all diseases in the epidemiology table."""

    if approval_prob_lookup is None:
        approval_prob_lookup = {}

    # Index drugs and pipeline by disease
    drugs_by_disease = defaultdict(list)
    for d in drugs_data:
        drugs_by_disease[d.get("disease_name", "")].append(d)

    pipeline_by_disease = defaultdict(list)
    for p in pipeline_data:
        pipeline_by_disease[p.get("disease_name", "")].append(p)

    science_by_disease = defaultdict(list)
    for s in science_data:
        science_by_disease[s.get("disease_name", "")].append(s)

    scores = []

    for disease in epi_data:
        name = disease.get("disease_name", "Unknown")
        area = disease.get("therapeutic_area", "Unknown")
        specialty_segment = disease.get("specialty_segment", "Unknown")
        prevalence = safe_int(disease.get("us_prevalence"), 0)
        severity = disease.get("severity_category", "chronic-manageable")
        mortality = safe_float(disease.get("mortality_rate_per_100k"), 0)

        drugs = drugs_by_disease.get(name, [])
        pipeline = pipeline_by_disease.get(name, [])
        science = science_by_disease.get(name, [])

        approved_count = len(drugs)

        # Weighted pipeline count
        active_pipeline = [p for p in pipeline if p.get("trial_status") in
                          ("RECRUITING", "NOT_YET_RECRUITING", "ACTIVE_NOT_RECRUITING", None, "")]
        pipeline_count = len(active_pipeline)
        weighted_pipeline = sum(
            PHASE_WEIGHTS.get(p.get("trial_phase", "PHASE1"), 0.1)
            for p in active_pipeline
        )

        recent_pubs = len(science)

        # ── Unmet Patient Need ────────────────────────────────────
        prevalence_score = normalize_log(max(prevalence, 1), 1000, 50_000_000) * 10
        severity_weight = SEVERITY_WEIGHTS.get(severity, 0.5)
        treatment_adequacy = min(approved_count / 5, 1.0)
        mortality_factor = normalize_log(max(mortality, 0.01), 0.1, 100) * 10 if mortality > 0 else 0

        unmet_need = (
            prevalence_score * 0.30 +
            severity_weight * 10 * 0.35 +
            (1 - treatment_adequacy) * 10 * 0.25 +
            mortality_factor * 0.10
        )
        unmet_need = clamp(unmet_need)

        # ── Drug Coverage Gap ─────────────────────────────────────
        total_coverage = approved_count + weighted_pipeline
        if prevalence > 0 and total_coverage > 0:
            coverage_per_million = total_coverage / (prevalence / 1_000_000)
            raw_gap = 10 - normalize_log(max(coverage_per_million, 0.01), 0.01, 1000) * 10
        elif total_coverage == 0:
            raw_gap = 10.0
        else:
            raw_gap = 5.0

        # MOA diversity penalty
        pipeline_moas = set(p.get("mechanism_of_action", "") for p in active_pipeline
                           if p.get("mechanism_of_action"))
        unique_moas = len(pipeline_moas)
        moa_diversity = min(unique_moas / 3, 1.0) if unique_moas > 0 else 0
        diversity_adj = (1 - moa_diversity) * 1.5

        coverage_gap = clamp(raw_gap + diversity_adj)

        # ── Scientific Complexity ─────────────────────────────────
        terminated = len([p for p in pipeline if p.get("trial_status") in ("TERMINATED", "WITHDRAWN")])
        total_trials = len(pipeline)
        if total_trials > 0:
            failure_ratio = terminated / total_trials
        else:
            failure_ratio = 0.5  # unknown = moderate complexity

        failure_score = failure_ratio * 10

        # Novelty from approved drugs
        all_drug_moas = drugs + active_pipeline
        novel_count = sum(1 for d in all_drug_moas
                         if d.get("moa_novelty") in ("novel", "emerging"))
        total_moa = len(all_drug_moas)
        novelty_fraction = novel_count / total_moa if total_moa > 0 else 0.7
        novelty_score = novelty_fraction * 10

        # Modality diversity
        routes = set(d.get("route_of_admin", "") for d in drugs if d.get("route_of_admin"))
        modality_score = min(len(routes) / 4, 1.0) * 10 if routes else 5.0

        # Research intensity
        research_intensity = normalize_log(max(recent_pubs, 1), 1, 10000) * 10

        complexity = (
            failure_score * 0.35 +
            novelty_score * 0.25 +
            modality_score * 0.15 +
            research_intensity * 0.25
        )
        complexity = clamp(complexity)

        # ── Regulatory Advantage Score (v2.0) ──────────────────────
        total_drugs = approved_count + pipeline_count
        if total_drugs == 0:
            regulatory_advantage = 5.0
        else:
            orphan_count = sum(1 for d in (drugs + pipeline) if safe_bool(d.get("orphan_drug", False)) or safe_bool(d.get("orphan_designation", False)))
            breakthrough_count = sum(1 for d in (drugs + pipeline) if safe_bool(d.get("breakthrough_therapy", False)) or safe_bool(d.get("breakthrough_designation", False)))
            fast_track_count = sum(1 for d in (drugs + pipeline) if safe_bool(d.get("fast_track", False)))

            orphan_fraction = orphan_count / total_drugs
            breakthrough_fraction = breakthrough_count / total_drugs
            fast_track_fraction = fast_track_count / total_drugs
            rare_bonus = 2.0 if prevalence < 200000 else 0.0

            regulatory_advantage = clamp(
                orphan_fraction * 10 * 0.35 +
                breakthrough_fraction * 10 * 0.35 +
                fast_track_fraction * 10 * 0.15 +
                rare_bonus * 0.15,
                0, 10)

        # ── Health Economics Score (v2.0) ──────────────────────────
        cost_per_patient = safe_float(disease.get("annual_cost_per_patient_usd"), 0)
        total_burden = safe_float(disease.get("total_us_economic_burden_usd"), 0)
        qaly = safe_float(disease.get("qaly_burden"), 0)

        if cost_per_patient > 0:
            cost_score = normalize_log(max(cost_per_patient, 1), 1000, 500000) * 10
        else:
            cost_score = 5.0

        if total_burden > 0:
            burden_score = normalize_log(max(total_burden, 1), 1_000_000, 500_000_000_000) * 10
        else:
            burden_score = 5.0

        if qaly > 0:
            qaly_score = normalize_log(max(qaly, 0.001), 0.01, 1.0) * 10
        else:
            qaly_score = 5.0

        gap_amplifier = 1.0 + (1 - min(approved_count / 5, 1.0)) * 0.3

        health_econ = clamp(
            (cost_score * 0.30 + burden_score * 0.40 + qaly_score * 0.30) * gap_amplifier,
            0, 10)

        # ── Approval Probability & Expected New Approvals (v2.0) ────
        approval_probs = approval_prob_lookup.get(area, {})
        expected_new_approvals = 0.0
        approval_probability_sum = 0.0
        approval_count = 0

        for drug in active_pipeline:
            phase = drug.get("trial_phase", "PHASE1")
            base_prob = approval_probs.get(phase, 0.5)

            # Apply orphan and breakthrough multipliers
            if safe_bool(drug.get("orphan_drug", False)) or safe_bool(drug.get("orphan_designation", False)):
                base_prob *= 1.25

            if safe_bool(drug.get("breakthrough_therapy", False)) or safe_bool(drug.get("breakthrough_designation", False)):
                base_prob *= 1.20

            # Apply novel MOA penalty
            if drug.get("moa_novelty", "") == "novel":
                base_prob *= 0.70

            # Phase-adjusted probability for expected approvals
            phase_weight = APPROVAL_PHASE_WEIGHTS.get(phase, 0.1)
            expected_new_approvals += base_prob * phase_weight

            approval_probability_sum += base_prob
            approval_count += 1

        if approval_count > 0:
            approval_probability_avg = clamp(approval_probability_sum / approval_count * 10)
        else:
            approval_probability_avg = 5.0

        expected_new_approvals = round(expected_new_approvals, 2)

        # ── v2.0 Composite Score ──────────────────────────────────
        composite = (
            unmet_need * W_UNMET_NEED +
            coverage_gap * W_COVERAGE_GAP +
            (10 - complexity) * W_COMPLEXITY_INV +
            regulatory_advantage * W_REGULATORY_ADV +
            health_econ * W_HEALTH_ECON
        )
        composite = clamp(composite)

        # Underserved ratio
        denominator = approved_count + weighted_pipeline
        underserved_ratio = prevalence / denominator if denominator > 0 else (
            prevalence if prevalence > 0 else 0
        )

        scores.append({
            "disease_name": name,
            "therapeutic_area": area,
            "specialty_segment": specialty_segment,
            "us_prevalence": prevalence,
            "approved_drug_count": approved_count,
            "pipeline_drug_count": pipeline_count,
            "weighted_pipeline_count": round(weighted_pipeline, 1),
            "approval_probability_avg": round(approval_probability_avg, 1),
            "expected_new_approvals": expected_new_approvals,
            "recent_publications": recent_pubs,
            "regulatory_advantage_score": round(regulatory_advantage, 1),
            "health_econ_score": round(health_econ, 1),
            "unmet_need_score": round(unmet_need, 1),
            "drug_coverage_gap": round(coverage_gap, 1),
            "scientific_complexity": round(complexity, 1),
            "composite_whitespace_score": round(composite, 1),
            "underserved_ratio": round(underserved_ratio, 1),
            "score_date": datetime.now().strftime("%Y-%m-%d"),
        })

    # Sort by composite score descending
    scores.sort(key=lambda x: x["composite_whitespace_score"], reverse=True)
    return scores


def write_scores(wb, scores):
    """Write scores to the White Space Scores tab, replacing existing data."""
    ws = wb["White Space Scores"]

    # Clear existing data (keep headers)
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # v2.0 Column order (includes new health econ, regulatory, approval probability columns)
    columns = [
        "disease_name", "therapeutic_area", "specialty_segment", "us_prevalence",
        "approved_drug_count", "pipeline_drug_count", "weighted_pipeline_count",
        "approval_probability_avg", "expected_new_approvals", "recent_publications",
        "regulatory_advantage_score", "health_econ_score",
        "unmet_need_score", "drug_coverage_gap", "scientific_complexity",
        "composite_whitespace_score", "underserved_ratio", "score_date"
    ]

    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Color-code by composite score
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
    med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # red

    for row_idx, score in enumerate(scores, 2):
        composite = score.get("composite_whitespace_score", 0)
        if composite >= 7:
            fill = high_fill
        elif composite >= 4:
            fill = med_fill
        else:
            fill = low_fill

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = score.get(col_name, "")
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill

            # Number formatting
            if col_name in ("unmet_need_score", "drug_coverage_gap", "scientific_complexity",
                            "composite_whitespace_score", "regulatory_advantage_score",
                            "health_econ_score", "approval_probability_avg"):
                cell.number_format = "0.0"
            elif col_name == "underserved_ratio":
                cell.number_format = "#,##0.0"
            elif col_name in ("weighted_pipeline_count",):
                cell.number_format = "0.0"
            elif col_name == "expected_new_approvals":
                cell.number_format = "0.00"
            elif col_name in ("us_prevalence",):
                cell.number_format = "#,##0"


def append_historical_snapshot(wb, scores):
    """Append a snapshot of current scores to the Historical Trends tab."""
    try:
        ws = wb["Historical Trends"]
    except KeyError:
        # Create the tab if it doesn't exist
        ws = wb.create_sheet("Historical Trends")
        ws.append(["snapshot_date", "disease_name", "therapeutic_area", "composite_score",
                   "unmet_need_score", "drug_coverage_gap", "regulatory_advantage_score",
                   "health_econ_score"])

    snapshot_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for score in scores:
        ws.append([
            snapshot_date,
            score.get("disease_name", ""),
            score.get("therapeutic_area", ""),
            score.get("composite_whitespace_score", ""),
            score.get("unmet_need_score", ""),
            score.get("drug_coverage_gap", ""),
            score.get("regulatory_advantage_score", ""),
            score.get("health_econ_score", ""),
        ])


def main():
    parser = argparse.ArgumentParser(description="Score diseases for white space analysis (v2.0)")
    parser.add_argument("workbook_path", help="Path to the Excel workbook")
    args = parser.parse_args()

    if not os.path.exists(args.workbook_path):
        print(f"Error: Workbook not found: {args.workbook_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.workbook_path)

    # Read all required tabs
    epi_data = read_tab_as_dicts(wb["Disease Epidemiology"])
    drugs_data = read_tab_as_dicts(wb["Approved Drugs"])
    pipeline_data = read_tab_as_dicts(wb["Pipeline Drugs"])
    science_data = read_tab_as_dicts(wb["Basic Science Activity"])

    if not epi_data:
        print("No diseases found in epidemiology tab. Add diseases first.")
        sys.exit(1)

    # Load approval probability data (optional)
    approval_prob_lookup = build_approval_probability_lookup(wb)

    print(f"Scoring {len(epi_data)} diseases (v2.0 with regulatory advantage & health economics)...")
    scores = compute_scores(epi_data, drugs_data, pipeline_data, science_data, approval_prob_lookup)

    write_scores(wb, scores)
    append_historical_snapshot(wb, scores)
    wb.save(args.workbook_path)

    # Print top 20 with new v2.0 dimensions
    print(f"\n{'='*110}")
    print(f"TOP WHITE SPACE OPPORTUNITIES (v2.0 - Composite Score Ranking)")
    print(f"{'='*110}")
    print(f"{'Rank':<5} {'Disease':<28} {'Unmet':>6} {'Gap':>6} {'RegAd':>6} {'HealthE':>7} {'Score':>7}")
    print(f"{'-'*110}")

    for i, s in enumerate(scores[:20], 1):
        disease_name = s['disease_name'][:27]
        print(f"{i:<5} {disease_name:<28} "
              f"{s['unmet_need_score']:>6.1f} "
              f"{s['drug_coverage_gap']:>6.1f} "
              f"{s['regulatory_advantage_score']:>6.1f} "
              f"{s['health_econ_score']:>7.1f} "
              f"{s['composite_whitespace_score']:>7.1f}")

    print(f"\nScores written to:")
    print(f"  - White Space Scores tab (v2.0 scoring with new columns)")
    print(f"  - Historical Trends tab (snapshot appended)")


if __name__ == "__main__":
    main()

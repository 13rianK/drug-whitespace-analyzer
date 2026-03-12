#!/usr/bin/env python3
"""
Build or update the Drug Whitespace Database Excel workbook.

Usage:
    python build_workbook.py <path_to_workbook>
    python build_workbook.py <path_to_workbook> --add-diseases <json_file>

The JSON file for --add-diseases should have the structure:
{
    "epidemiology": [...],
    "approved_drugs": [...],
    "pipeline_drugs": [...],
    "basic_science": [...]
}

Each array contains dicts matching the column schema in references/data_schema.md.
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Column Definitions ────────────────────────────────────────────────

TABS = {
    "Disease Epidemiology": [
        "disease_name", "icd10_code", "icd10_chapter", "icd10_chapter_name",
        "therapeutic_area", "us_prevalence",
        "global_prevalence", "us_annual_incidence", "mortality_rate_per_100k",
        "severity_category", "primary_determinants", "specialty_segment",
        "age_group_affected", "annual_cost_per_patient_usd",
        "total_us_economic_burden_usd", "qaly_burden", "data_source",
        "source_url", "date_pulled"
    ],
    "Approved Drugs": [
        "disease_name", "drug_brand_name", "drug_generic_name",
        "mechanism_of_action", "moa_novelty", "approval_year",
        "manufacturer", "route_of_admin", "orphan_drug", "breakthrough_therapy",
        "accelerated_approval", "annual_cost_usd", "data_source", "date_pulled"
    ],
    "Pipeline Drugs": [
        "disease_name", "nct_id", "intervention_name", "trial_phase",
        "trial_status", "sponsor", "enrollment_target",
        "mechanism_of_action", "orphan_drug", "breakthrough_designation",
        "fast_track", "estimated_completion", "start_date", "date_pulled"
    ],
    "Basic Science Activity": [
        "disease_name", "activity_type", "title", "source_org",
        "description", "funding_amount_usd", "pmid", "doi",
        "publication_date", "date_pulled"
    ],
    "White Space Scores": [
        "disease_name", "therapeutic_area", "specialty_segment", "us_prevalence",
        "approved_drug_count", "pipeline_drug_count", "weighted_pipeline_count",
        "approval_probability_avg", "expected_new_approvals",
        "recent_publications", "regulatory_advantage_score", "health_econ_score",
        "unmet_need_score", "drug_coverage_gap", "scientific_complexity",
        "composite_whitespace_score", "underserved_ratio", "score_date"
    ],
    "Historical Trends": [
        "disease_name", "snapshot_date", "approved_drug_count",
        "pipeline_drug_count", "composite_whitespace_score", "underserved_ratio",
        "us_prevalence", "notes"
    ],
    "Approval Probability": [
        "therapeutic_area", "phase_1_to_approval", "phase_2_to_approval",
        "phase_3_to_approval", "orphan_boost", "breakthrough_boost",
        "novel_moa_penalty", "data_source"
    ],
    "Metadata": [
        "parameter", "value"
    ]
}

# ── Styling ───────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Columns that should be formatted as numbers
NUMBER_COLS = {
    "us_prevalence", "global_prevalence", "us_annual_incidence",
    "mortality_rate_per_100k", "approval_year", "enrollment_target",
    "approved_drug_count", "pipeline_drug_count", "weighted_pipeline_count",
    "recent_publications", "unmet_need_score", "drug_coverage_gap",
    "scientific_complexity", "composite_whitespace_score", "underserved_ratio",
    "annual_cost_per_patient_usd", "total_us_economic_burden_usd", "qaly_burden",
    "annual_cost_usd", "funding_amount_usd", "approval_probability_avg",
    "expected_new_approvals", "regulatory_advantage_score", "health_econ_score",
    "phase_1_to_approval", "phase_2_to_approval", "phase_3_to_approval",
    "orphan_boost", "breakthrough_boost", "novel_moa_penalty"
}


# ── ICD-10 Chapter Lookup ─────────────────────────────────────────────

_ICD10_MAP = {
    "A": ("I", "Infectious & Parasitic Diseases"),
    "B": ("I", "Infectious & Parasitic Diseases"),
    "C": ("II", "Neoplasms"),
    "D0": ("II", "Neoplasms"), "D1": ("II", "Neoplasms"), "D2": ("II", "Neoplasms"),
    "D3": ("II", "Neoplasms"), "D4": ("II", "Neoplasms"),
    "D5": ("III", "Blood & Immune Disorders"), "D6": ("III", "Blood & Immune Disorders"),
    "D7": ("III", "Blood & Immune Disorders"), "D8": ("III", "Blood & Immune Disorders"),
    "D9": ("III", "Blood & Immune Disorders"),
    "E": ("IV", "Endocrine, Nutritional & Metabolic"),
    "F": ("V", "Mental & Behavioral Disorders"),
    "G": ("VI", "Nervous System"),
    "H0": ("VII", "Eye & Adnexa"), "H1": ("VII", "Eye & Adnexa"),
    "H2": ("VII", "Eye & Adnexa"), "H3": ("VII", "Eye & Adnexa"),
    "H4": ("VII", "Eye & Adnexa"), "H5": ("VII", "Eye & Adnexa"),
    "H6": ("VIII", "Ear & Mastoid"), "H7": ("VIII", "Ear & Mastoid"),
    "H8": ("VIII", "Ear & Mastoid"), "H9": ("VIII", "Ear & Mastoid"),
    "I": ("IX", "Circulatory System"),
    "J": ("X", "Respiratory System"),
    "K": ("XI", "Digestive System"),
    "L": ("XII", "Skin & Subcutaneous Tissue"),
    "M": ("XIII", "Musculoskeletal & Connective Tissue"),
    "N": ("XIV", "Genitourinary System"),
    "O": ("XV", "Pregnancy & Childbirth"),
    "P": ("XVI", "Perinatal Conditions"),
    "Q": ("XVII", "Congenital Malformations"),
    "R": ("XVIII", "Symptoms & Signs"),
    "S": ("XIX", "Injury & Poisoning"), "T": ("XIX", "Injury & Poisoning"),
    "V": ("XX", "External Causes"), "W": ("XX", "External Causes"),
    "X": ("XX", "External Causes"), "Y": ("XX", "External Causes"),
    "Z": ("XXI", "Factors Influencing Health"),
}

def icd10_chapter_lookup(code):
    """Return (chapter_number, chapter_name) from an ICD-10 code."""
    if not code:
        return ("", "")
    code = str(code).strip().upper().split(",")[0].strip()
    if len(code) >= 2 and code[:2] in _ICD10_MAP:
        return _ICD10_MAP[code[:2]]
    if code[0] in _ICD10_MAP:
        return _ICD10_MAP[code[0]]
    return ("", "")


def create_workbook(path):
    """Create a new workbook with all tabs and headers."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    for tab_name, columns in TABS.items():
        ws = wb.create_sheet(title=tab_name)

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            # Human-readable header
            cell.value = col_name.replace("_", " ").title()
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Set column widths
        for col_idx, col_name in enumerate(columns, 1):
            width = max(len(col_name) + 4, 15)
            if col_name in ("description", "title", "source_url", "data_source", "notes"):
                width = 40
            elif col_name in ("disease_name", "mechanism_of_action", "therapeutic_area",
                              "age_group_affected", "specialty_segment"):
                width = 25
            elif col_name in ("intervention_name", "drug_brand_name", "drug_generic_name"):
                width = 30
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if columns:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Populate metadata
    ws_meta = wb["Metadata"]
    metadata_rows = [
        ("database_created", datetime.now().isoformat()),
        ("last_full_update", datetime.now().isoformat()),
        ("total_diseases", "0"),
        ("total_approved_drugs", "0"),
        ("total_pipeline_drugs", "0"),
        ("schema_version", "2.0"),
        ("scoring_version", "2.0"),
        ("notes", "Initial creation"),
    ]
    for row_idx, (param, value) in enumerate(metadata_rows, 2):
        ws_meta.cell(row=row_idx, column=1, value=param).font = DATA_FONT
        ws_meta.cell(row=row_idx, column=2, value=value).font = DATA_FONT

    # Pre-populate Approval Probability tab with reference data
    ws_approval = wb["Approval Probability"]
    approval_data = [
        ("Oncology", 0.052, 0.155, 0.592, 1.40, 1.25, 0.70),
        ("Neurology", 0.061, 0.132, 0.548, 1.40, 1.25, 0.70),
        ("Cardiovascular", 0.083, 0.181, 0.625, 1.40, 1.25, 0.70),
        ("Metabolic", 0.091, 0.203, 0.651, 1.40, 1.25, 0.70),
        ("Endocrine/Metabolic", 0.091, 0.203, 0.651, 1.40, 1.25, 0.70),
        ("Infectious Disease", 0.115, 0.228, 0.673, 1.40, 1.25, 0.70),
        ("Respiratory", 0.072, 0.167, 0.589, 1.40, 1.25, 0.70),
        ("Autoimmune", 0.078, 0.172, 0.561, 1.40, 1.25, 0.70),
        ("Hematology", 0.102, 0.215, 0.648, 1.40, 1.25, 0.70),
        ("Rare Disease", 0.125, 0.253, 0.687, 1.0, 1.25, 0.70),
        ("Mental Health", 0.058, 0.121, 0.512, 1.40, 1.25, 0.70),
        ("Other", 0.08, 0.17, 0.60, 1.40, 1.25, 0.70),
    ]
    data_source = "BIO/QLS 2024 Industry Report"

    for row_idx, approval_row in enumerate(approval_data, 2):
        therapeutic_area, p1, p2, p3, orphan, breakthrough, novel_moa = approval_row
        ws_approval.cell(row=row_idx, column=1, value=therapeutic_area).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=2, value=p1).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=3, value=p2).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=4, value=p3).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=5, value=orphan).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=6, value=breakthrough).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=7, value=novel_moa).font = DATA_FONT
        ws_approval.cell(row=row_idx, column=8, value=data_source).font = DATA_FONT

        # Format numeric columns
        for col_idx in [2, 3, 4, 5, 6, 7]:
            ws_approval.cell(row=row_idx, column=col_idx).number_format = "0.00"

    wb.save(path)
    print(f"Created new workbook: {path}")
    return wb


def add_data(wb, tab_name, rows):
    """Append rows to a tab. Each row is a dict with column keys."""
    if tab_name not in TABS:
        print(f"Warning: Unknown tab '{tab_name}', skipping")
        return 0

    ws = wb[tab_name]
    columns = TABS[tab_name]
    start_row = ws.max_row + 1
    added = 0

    for row_data in rows:
        # Auto-populate ICD-10 chapter if missing
        if tab_name == "Disease Epidemiology" and "icd10_code" in row_data:
            if not row_data.get("icd10_chapter"):
                ch_num, ch_name = icd10_chapter_lookup(row_data.get("icd10_code", ""))
                row_data["icd10_chapter"] = ch_num
                row_data["icd10_chapter_name"] = ch_name

        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=start_row + added, column=col_idx)

            # Type coercion for numeric columns
            if col_name in NUMBER_COLS and value != "" and value is not None:
                try:
                    value = float(value)
                    if value == int(value) and col_name not in (
                        "mortality_rate_per_100k", "unmet_need_score",
                        "drug_coverage_gap", "scientific_complexity",
                        "composite_whitespace_score", "underserved_ratio",
                        "weighted_pipeline_count"
                    ):
                        value = int(value)
                except (ValueError, TypeError):
                    pass

            cell.value = value
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Number formatting
            if col_name in ("unmet_need_score", "drug_coverage_gap",
                            "scientific_complexity", "composite_whitespace_score",
                            "approval_probability_avg", "regulatory_advantage_score",
                            "health_econ_score"):
                cell.number_format = "0.0"
            elif col_name in ("underserved_ratio", "qaly_burden"):
                cell.number_format = "#,##0.0"
            elif col_name in ("weighted_pipeline_count", "phase_1_to_approval",
                              "phase_2_to_approval", "phase_3_to_approval",
                              "orphan_boost", "breakthrough_boost", "novel_moa_penalty"):
                cell.number_format = "0.0"
            elif col_name in ("us_prevalence", "global_prevalence",
                              "us_annual_incidence", "enrollment_target",
                              "annual_cost_per_patient_usd", "total_us_economic_burden_usd",
                              "annual_cost_usd", "funding_amount_usd",
                              "expected_new_approvals"):
                cell.number_format = "#,##0"
            elif col_name == "mortality_rate_per_100k":
                cell.number_format = "0.0"

        added += 1

    print(f"Added {added} rows to '{tab_name}'")
    return added


def update_metadata(wb, key, value):
    """Update a metadata parameter."""
    ws = wb["Metadata"]
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[0].value == key:
            row[1].value = str(value)
            return
    # If not found, append
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=key).font = DATA_FONT
    ws.cell(row=next_row, column=2, value=str(value)).font = DATA_FONT


def count_rows(wb, tab_name):
    """Count data rows (excluding header) in a tab."""
    ws = wb[tab_name]
    count = 0
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value is not None and row[0].value != "":
            count += 1
    return count


def main():
    parser = argparse.ArgumentParser(description="Build or update drug whitespace database")
    parser.add_argument("workbook_path", help="Path to the Excel workbook")
    parser.add_argument("--add-diseases", help="JSON file with disease data to add")
    args = parser.parse_args()

    # Create or open workbook
    if os.path.exists(args.workbook_path):
        print(f"Opening existing workbook: {args.workbook_path}")
        wb = openpyxl.load_workbook(args.workbook_path)
    else:
        wb = create_workbook(args.workbook_path)

    # Add data if provided
    if args.add_diseases:
        with open(args.add_diseases, "r") as f:
            data = json.load(f)

        tab_mapping = {
            "epidemiology": "Disease Epidemiology",
            "approved_drugs": "Approved Drugs",
            "pipeline_drugs": "Pipeline Drugs",
            "basic_science": "Basic Science Activity",
            "scores": "White Space Scores",
        }

        for key, tab_name in tab_mapping.items():
            if key in data and data[key]:
                add_data(wb, tab_name, data[key])

        # Update metadata counts
        update_metadata(wb, "total_diseases", count_rows(wb, "Disease Epidemiology"))
        update_metadata(wb, "total_approved_drugs", count_rows(wb, "Approved Drugs"))
        update_metadata(wb, "total_pipeline_drugs", count_rows(wb, "Pipeline Drugs"))
        update_metadata(wb, "last_full_update", datetime.now().isoformat())

    wb.save(args.workbook_path)
    print(f"Saved workbook: {args.workbook_path}")

    # Print summary
    print(f"\nDatabase Summary:")
    print(f"  Diseases: {count_rows(wb, 'Disease Epidemiology')}")
    print(f"  Approved Drugs: {count_rows(wb, 'Approved Drugs')}")
    print(f"  Pipeline Drugs: {count_rows(wb, 'Pipeline Drugs')}")
    print(f"  Basic Science Entries: {count_rows(wb, 'Basic Science Activity')}")
    print(f"  Scored Diseases: {count_rows(wb, 'White Space Scores')}")


if __name__ == "__main__":
    main()
